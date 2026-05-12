#!/usr/bin/env python3
"""Export Experiment 1.2 sessions to Excel.

One sheet per session. Each sheet: rows = rounds, columns = P1..P4, mean, pot, payoff_each.
A summary sheet at the front shows R1, Rlast, mean, decay slope for every session.

Usage:
    python scripts/sessions_to_excel.py
    python scripts/sessions_to_excel.py --input outputs/exp1_2_pilot --output outputs/exp1_2_results.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

# ── palette ──────────────────────────────────────────────────────────────────
_COL_HEADER   = "1F4E79"   # dark blue
_COL_SUBHEAD  = "2E75B6"   # mid blue
_COL_DEFECT   = "FCE4D6"   # light red — contribution == 0
_COL_FULL     = "E2EFDA"   # light green — contribution == endowment
_COL_STRIPE   = "F2F2F2"   # light grey row stripe
_WHITE        = "FFFFFF"
_YELLOW       = "FFF2CC"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _hdr_font(white: bool = True) -> Font:
    return Font(bold=True, color=_WHITE if white else "000000")


def _safe_slope(xs: List[float], ys: List[float]) -> Optional[float]:
    n = len(xs)
    if n < 2:
        return None
    mx = sum(xs) / n
    my = sum(ys) / n
    denom = sum((x - mx) ** 2 for x in xs)
    if denom == 0:
        return 0.0
    return sum((xs[i] - mx) * (ys[i] - my) for i in range(n)) / denom


def _col_width(ws, col_idx: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_session_sheet(
    wb: openpyxl.Workbook,
    session: Dict[str, Any],
    skip_incomplete_rounds: bool = True,
) -> Dict[str, Any]:
    model     = session["model_requested"]
    sess_idx  = session["session"]
    temp      = session["temperature"]
    transp    = session["transparency"]
    endow     = session["endowment"]
    mult      = session["multiplier"]
    n_players = session["num_players"]

    # sheet name: shortened model + session
    short = model.split("/")[-1][:20]
    sheet_name = f"{short}_s{sess_idx}"

    ws = wb.create_sheet(title=sheet_name)

    # ── title row ────────────────────────────────────────────────────────────
    title = f"{model}  |  session {sess_idx}  |  T={temp}  |  transparency={transp}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + n_players + 3)
    cell = ws.cell(1, 1, title)
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 18

    # ── column headers ───────────────────────────────────────────────────────
    headers = ["Round"] + [f"P{p}" for p in range(1, n_players + 1)] + ["mean", "pot", "payoff_each"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font = _hdr_font()
        c.fill = _fill(_COL_HEADER)
        c.alignment = Alignment(horizontal="center")

    _col_width(ws, 1, 7)
    for i in range(2, n_players + 2):
        _col_width(ws, i, 8)
    _col_width(ws, n_players + 2, 8)   # mean
    _col_width(ws, n_players + 3, 8)   # pot
    _col_width(ws, n_players + 4, 12)  # payoff_each

    # ── data rows ────────────────────────────────────────────────────────────
    rounds_data = session.get("rounds", [])
    round_means: List[float] = []
    data_row = 3

    for rnd in rounds_data:
        if skip_incomplete_rounds and rnd.get("incomplete"):
            continue

        round_num = rnd["round"]
        actions   = rnd.get("actions") or []
        payoffs   = rnd.get("payoffs") or []
        acts_pad  = list(actions) + [None] * (n_players - len(actions))

        stripe = _fill(_COL_STRIPE) if round_num % 2 == 0 else _fill(_WHITE)

        # Round number cell
        c = ws.cell(data_row, 1, round_num)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.fill = _fill(_COL_SUBHEAD)
        c.font = _hdr_font()

        # Player contribution cells
        for p_idx, act in enumerate(acts_pad):
            c = ws.cell(data_row, p_idx + 2, act)
            c.alignment = Alignment(horizontal="center")
            c.number_format = "0.0"
            if act is None:
                c.fill = stripe
            elif act == 0:
                c.fill = _fill(_COL_DEFECT)
            elif act == endow:
                c.fill = _fill(_COL_FULL)
            else:
                c.fill = stripe

        parsed = [a for a in actions if isinstance(a, (int, float))]
        mean_val = sum(parsed) / len(parsed) if parsed else None
        pot_val  = sum(parsed) * mult if parsed else None
        payoff_avg = sum(payoffs) / len(payoffs) if payoffs else None

        if mean_val is not None:
            round_means.append(mean_val)

        for col, val, fmt in [
            (n_players + 2, mean_val, "0.00"),
            (n_players + 3, pot_val,  "0.0"),
            (n_players + 4, payoff_avg, "0.00"),
        ]:
            c = ws.cell(data_row, col, val)
            c.alignment = Alignment(horizontal="center")
            c.number_format = fmt
            c.fill = stripe

        data_row += 1

    # ── footer: session stats ────────────────────────────────────────────────
    data_row += 1
    labels_row = data_row
    stats: Dict[str, Any] = {}
    if round_means:
        xs = list(range(1, len(round_means) + 1))
        slope = _safe_slope(xs, round_means)
        stats = {
            "avg_mean":  sum(round_means) / len(round_means),
            "R1_mean":   round_means[0],
            "Rlast_mean": round_means[-1],
            "delta":     round_means[0] - round_means[-1],
            "slope":     slope,
        }
        footer = [
            ("session avg", stats["avg_mean"],  "0.00"),
            ("R1 mean",     stats["R1_mean"],   "0.00"),
            ("Rlast mean",  stats["Rlast_mean"],"0.00"),
            ("ΔR1-Rlast",   stats["delta"],     "+0.00;-0.00;0.00"),
            ("OLS slope",   stats["slope"],     "+0.0000;-0.0000;0.0000"),
        ]
        for label, val, fmt in footer:
            lc = ws.cell(labels_row, 1, label)
            lc.font = Font(bold=True)
            lc.fill = _fill(_YELLOW)
            vc = ws.cell(labels_row, 2, val)
            vc.number_format = fmt
            vc.fill = _fill(_YELLOW)
            labels_row += 1

    return {"sheet": sheet_name, "model": model, "session": sess_idx, **stats}


def _write_summary_sheet(wb: openpyxl.Workbook, summaries: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet(title="Summary", index=0)

    headers = ["Model", "Session", "Avg mean", "R1 mean", "Rlast mean", "ΔR1-Rlast", "OLS slope", "Sheet"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = _hdr_font()
        c.fill = _fill(_COL_HEADER)
        c.alignment = Alignment(horizontal="center")

    widths = [36, 9, 10, 10, 12, 11, 11, 28]
    for i, w in enumerate(widths, 1):
        _col_width(ws, i, w)

    for row_idx, s in enumerate(summaries, 2):
        stripe = _fill(_COL_STRIPE) if row_idx % 2 == 0 else _fill(_WHITE)
        values = [
            s.get("model", ""),
            s.get("session", ""),
            s.get("avg_mean"),
            s.get("R1_mean"),
            s.get("Rlast_mean"),
            s.get("delta"),
            s.get("slope"),
            s.get("sheet", ""),
        ]
        fmts = ["@", "0", "0.00", "0.00", "0.00",
                "+0.00;-0.00;0.00", "+0.0000;-0.0000;0.0000", "@"]
        for col, (val, fmt) in enumerate(zip(values, fmts), 1):
            c = ws.cell(row_idx, col, val)
            c.number_format = fmt
            c.fill = stripe
            c.alignment = Alignment(horizontal="center" if col > 1 else "left")

            # colour the delta cell
            if col == 6 and isinstance(val, float):
                if val > 2:
                    c.fill = _fill(_COL_DEFECT)   # red — cooperation decayed
                elif val < -2:
                    c.fill = _fill(_COL_FULL)      # green — cooperation grew


def _iter_sessions(root: Path) -> List[Dict[str, Any]]:
    sessions = []
    skip = ("failed_attempt",)
    for sp in sorted(root.rglob("session_summary.json")):
        if any(s in str(sp) for s in skip):
            continue
        data = json.loads(sp.read_text())
        if data.get("incomplete"):
            continue
        sessions.append(data)
    return sessions


def main() -> None:
    parser = argparse.ArgumentParser(description="Export Exp 1.2 sessions to Excel.")
    parser.add_argument("--input",  type=Path, default=Path("outputs/exp1_2_pilot"))
    parser.add_argument("--output", type=Path, default=Path("outputs/exp1_2_results.xlsx"))
    args = parser.parse_args()

    sessions = _iter_sessions(args.input)
    if not sessions:
        print(f"No complete sessions found under {args.input}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheet_summaries = []
    for sess in sessions:
        stats = _write_session_sheet(wb, sess)
        sheet_summaries.append(stats)
        print(f"  {stats['model']:<35} session {stats['session']}  →  sheet '{stats['sheet']}'")

    _write_summary_sheet(wb, sheet_summaries)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"\nSaved {args.output}  ({len(sessions)} sessions, {len(sessions) + 1} sheets)")


if __name__ == "__main__":
    main()
